from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

from src.excel_export import export_excel_workbook


def test_excel_export_integrity():
    tmp_path = Path("outputs/debug/pytest_tmp_excel")
    tmp_path.mkdir(parents=True, exist_ok=True)

    datasets = {
        "Universe": pd.DataFrame({"company_name": ["A"], "ticker": ["AAA"]}),
        "Universe_Review": pd.DataFrame({"company_name": ["A"], "bucket_classification": ["primary"]}),
        "Hormuz_Ranking": pd.DataFrame({"ticker": ["AAA"], "combined_exposure_pct": [12.3]}),
        "Route_Risks": pd.DataFrame({"ticker": ["AAA"], "hormuz_used": [True]}),
        "Crude_Tracker": pd.DataFrame({"date": ["2026-01-02"], "brent_price": [80.0]}),
        "Fuel_Tracker": pd.DataFrame({"date": ["2026-01-02"], "ticker": ["AAA"], "blended_combined_fuels_price": [95.0]}),
        "Fuel_Geography_Weights": pd.DataFrame({"company": ["A"], "geography": ["US"], "weight": [1.0]}),
        "Equity_Tracker": pd.DataFrame({"date": ["2026-01-02"], "ticker": ["AAA"], "share_price": [50.0]}),
        "Operating_Mix": pd.DataFrame({"ticker": ["AAA"], "upstream_share_pct": [60.0], "downstream_share_pct": [40.0]}),
        "Earnings": pd.DataFrame({"ticker": ["AAA"], "next_earnings_date": ["2026-05-01"]}),
        "Assumptions": pd.DataFrame({"company": ["A"], "field_name": ["x"]}),
        "Source_Log": pd.DataFrame({"company": ["A"], "source_url": ["https://example.com"]}),
        "Missing_Data_Log": pd.DataFrame({"company": ["A"], "field_name": ["y"]}),
    }

    img = tmp_path / "chart.png"
    fig, ax = plt.subplots()
    ax.plot([1, 2], [1, 2])
    fig.savefig(img)
    plt.close(fig)

    chart_sheets = [
        {
            "sheet_name": "Chart_Test",
            "image_path": str(img),
            "explanation": "Test chart",
            "takeaway": "Works",
        }
    ]

    out_file = tmp_path / "analysis.xlsx"
    export_excel_workbook(
        output_path=out_file,
        datasets=datasets,
        chart_sheets=chart_sheets,
        methodology_summary="test",
    )

    assert out_file.exists()

    wb = openpyxl.load_workbook(out_file)
    assert "README" in wb.sheetnames
    assert "Universe" in wb.sheetnames
    assert "Chart_Test" in wb.sheetnames
